import openpyxl
import re
from datetime import datetime, date
from django.utils import timezone
from django.core.management.base import BaseCommand
from django.contrib.auth.hashers import make_password
from apps.modulo_admin.models import Usuario, VinculoProdutorRegional, VinculoTecnicoRegional, UF_Municipio
from django.contrib.auth.models import User

def validar_cpf(cpf):
    cpf = ''.join(filter(str.isdigit, cpf))
    if len(cpf) != 11:
        return False
    if cpf == cpf[0] * 11:
        return False
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    digito1 = 11 - (soma % 11)
    if digito1 >= 10:
        digito1 = 0
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    digito2 = 11 - (soma % 11)
    if digito2 >= 10:
        digito2 = 0
    return cpf[-2:] == f"{digito1}{digito2}"

def cpf_existe_na_base(cpf):
    return Usuario.objects.filter(cpf=cpf).exists()

def validar_data(data):
    if isinstance(data, (datetime, date)):
        return True
    if isinstance(data, str):
        try:
            datetime.strptime(data, '%Y-%m-%d %H:%M:%S')
            return True
        except ValueError:
            return False
    return False

def codigo_ibge_existe(ibge):
    return UF_Municipio.objects.filter(cod_ibge=ibge).exists()

def formato_celular(celular):
    padrao = re.compile(r'^\(\d{2}\) \d{5}-\d{4}$')
    return bool(padrao.match(celular))

def uf_sigla_existe(uf):
    return UF_Municipio.objects.filter(uf_sigla=uf).exists()

def criar_user(cpf, nome_completo, email):
    user = User()

    cpf_somente_numeros = ''.join(filter(str.isdigit, cpf))
    hashed_password = make_password(cpf_somente_numeros)

    user.username = cpf
    user.first_name = nome_completo
    user.email = email
    user.senha = hashed_password
    user.save()
    return user

def criar_usuario(user, cpf, nome_completo, data_nascimento, sexo, ibge, celular, email, tipo_usuario):
    usuario = Usuario()
    usuario.cpf = cpf
    usuario.nome_completo = nome_completo
    usuario.data_nascimento = data_nascimento
    usuario.user = user
    usuario.sexo = sexo
    usuario.ibge = ibge
    usuario.celular = celular
    usuario.email = email
    usuario.save()
    return usuario

def vincular_produtor(usuario, regional):
    vinculo = VinculoProdutorRegional()
    vinculo.usuario = usuario
    vinculo.regional = regional
    vinculo.data_inicio = date.today()
    vinculo.save()
    return vinculo

def vincular_tecnico(usuario, regional):
    vinculo = VinculoTecnicoRegional()
    vinculo.usuario = usuario
    vinculo.regional = regional
    vinculo.data_inicio = date.today()
    vinculo.save()
    return vinculo

def import_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2): 
        cpf = row[0].value
        nome_completo = row[1].value
        data_nascimento = row[2].value
        sexo = row[3].value
        ibge = row[4].value
        celular = row[5].value
        email = row[6].value
        tipo_usuario = row[7].value
        regional = row[8].value

        #Avaliar preenchimento dos campos
        campos_obrigatorios = [cpf, nome_completo, data_nascimento, sexo, ibge, celular, email, tipo_usuario, regional]
        campos_nomes = ['CPF', 'Nome Completo', 'Data de Nascimento', 'Sexo', 'Código IBGE do Município', 'Celular', 'Email', 'Tipo de Usuário', 'Regional']
        erros = [nome for valor, nome in zip(campos_obrigatorios, campos_nomes) if not valor]
        if erros:
            sheet.cell(row=row[0].row, column=10, value=f'Campo(s) obrigatório(s) não preenchido(s): {", ".join(erros)}')
            continue

        if not validar_cpf(cpf):
            sheet.cell(row=row[0].row, column=10, value="CPF inválido")
            continue

        if cpf_existe_na_base(cpf):
            sheet.cell(row=row[0].row, column=10, value="CPF já cadastrado")
            continue

        if not validar_data(str(data_nascimento)):
            sheet.cell(row=row[0].row, column=10, value='Data de nascimento inválida')
            continue

        if sexo not in ['Feminino', 'Masculino', 'Outro']:
            sheet.cell(row=row[0].row, column=10, value='Sexo inválido')
            continue

        if not codigo_ibge_existe(ibge):
            sheet.cell(row=row[0].row, column=10, value="Código IBGE do município não existe")
            continue

        if not formato_celular(celular):
            sheet.cell(row=row[0].row, column=10, value="Formato do celular inválido")
            continue

        if tipo_usuario not in ['Produtor', 'Técnico']:
            sheet.cell(row=row[0].row, column=10, value='Tipo de Usuário inválido')
            continue

        if not uf_sigla_existe(regional):
            sheet.cell(row=row[0].row, column=10, value='Regional inválida')
            continue

        try:
            user = criar_user(cpf=cpf, nome_completo=nome_completo, email=email)
            if user:
                usuario = criar_usuario(user, cpf, nome_completo, data_nascimento, sexo, ibge, celular, email, tipo_usuario)
                if tipo_usuario == 'Produtor':
                    try:
                        vincular_produtor(usuario=usuario, regional=regional)
                        sheet.cell(row=row[0].row, column=10, value="Produtor cadastrado com sucesso!")
                    except Exception as e:
                        sheet.cell(row=row[0].row, column=10, value="Erro na Vinculação do Produtor: " + str(e))
                elif tipo_usuario == 'Técnico':
                    try:
                        vincular_tecnico(usuario=usuario, regional=regional)
                        sheet.cell(row=row[0].row, column=10, value="Técnico cadastrado com sucesso!")
                    except Exception as e:
                        sheet.cell(row=row[0].row, column=10, value="Erro na Vinculação do Técnico: " + str(e))
        except Exception as e:
            sheet.cell(row=row[0].row, column=10, value="Erro no cadastro do Usuário: " + str(e))
            continue

    workbook.save(file_path)

class Command(BaseCommand):
    help = 'Importa lista de usuários de um arquivo Excel'

    def handle(self, *args, **kwargs):
        file_path = 'dados/lista_usuarios.xlsx'
        import_from_excel(file_path)
        self.stdout.write(self.style.SUCCESS('Lista de usuários processada!'))